import io
import json
import math
from datetime import datetime, timedelta
from pathlib import Path

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from sqlalchemy.orm import Session, joinedload

from ..config import UPLOAD_DIR, now as app_now
from ..database import get_db
from ..models import (
    Session as SessionModel, Staff, Assignment, Room,
    LTTalk, Category, SessionGroup, AppSetting,
)
from ..utils import upload_path

router = APIRouter(prefix="/api/export", tags=["export"])

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
HEADER_FILL_TEAL = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")
HEADER_FILL_BROWN = PatternFill(start_color="5D4037", end_color="5D4037", fill_type="solid")
HEADER_FILL_INDIGO = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

SESSION_CATS = ("general", "tech", "workshop", "keynote", "lt", "panel")
ROLE_LABELS_BASE = {"session": "セッション"}
CAT_LABELS_BASE = {
    "general": "一般", "tech": "技術", "workshop": "ワークショップ",
    "keynote": "基調講演", "lt": "LT", "panel": "パネルディスカッション", "overall": "全体",
}

PHOTO_PX = 48  # Excel内の写真サイズ (px)
ROW_HEIGHT_WITH_PHOTO = 45  # 写真付き行の高さ (pt)

SLOT_MINUTES = 5  # 全体スケジュールの時間スロット（分）


def _fmt(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%H:%M")


def _fmt_full(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y/%m/%d %H:%M")


def _fmt_md(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%m/%d %H:%M")


def _apply_header(ws, row, fill=HEADER_FILL):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _apply_border(ws, row):
    for cell in ws[row]:
        cell.border = THIN_BORDER
        cell.alignment = WRAP


def _auto_width(ws, max_rows=100):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for i, cell in enumerate(col):
            if i >= max_rows:
                break
            try:
                val = str(cell.value or "")
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _add_photo(ws, photo_path: str, col: int, row: int):
    """ワークシートのセルに元画像をそのまま埋め込む（セル内に収まるようリサイズ表示）"""
    file_path = upload_path(photo_path)
    if not file_path.exists():
        return
    try:
        img = PILImage.open(file_path)
        img.verify()  # 画像ファイルとして有効か確認
    except Exception:
        return
    xl_img = XlImage(str(file_path))
    xl_img.width = PHOTO_PX
    xl_img.height = PHOTO_PX
    cell_ref = f"{get_column_letter(col)}{row}"
    ws.add_image(xl_img, cell_ref)
    ws.row_dimensions[row].height = ROW_HEIGHT_WITH_PHOTO


def _safe_sheet_name(name: str) -> str:
    """Excelのシート名制約（31文字・禁止文字）に合わせて整形"""
    for ch in r"\/?*[]:":
        name = name.replace(ch, "-")
    return name[:31]


def _floor_slot(dt: datetime) -> datetime:
    return dt.replace(minute=dt.minute - dt.minute % SLOT_MINUTES, second=0, microsecond=0)


def _ceil_slot(dt: datetime) -> datetime:
    dt = dt.replace(second=0, microsecond=0)
    rem = dt.minute % SLOT_MINUTES
    if rem:
        dt += timedelta(minutes=SLOT_MINUTES - rem)
    return dt


@router.get("/excel")
def export_excel(db: Session = Depends(get_db)):
    """タブ構成に合わせてExcelファイルをエクスポート"""

    # データ取得
    rooms = db.query(Room).order_by(Room.id).all()
    sessions = (
        db.query(SessionModel)
        .options(
            joinedload(SessionModel.room),
            joinedload(SessionModel.lt_talks),
            joinedload(SessionModel.assignments).joinedload(Assignment.staff).joinedload(Staff.skills),
        )
        .order_by(SessionModel.start_time)
        .all()
    )
    staffs = (
        db.query(Staff)
        .options(
            joinedload(Staff.availabilities),
            joinedload(Staff.preferred_sessions),
            joinedload(Staff.assignments).joinedload(Assignment.session).joinedload(SessionModel.room),
        )
        .order_by(Staff.id)
        .all()
    )

    # 動的カテゴリ取得
    db_categories = db.query(Category).order_by(Category.order, Category.id).all()
    db_session_groups = db.query(SessionGroup).order_by(SessionGroup.order, SessionGroup.id).all()
    dynamic_cat_keys = [c.key for c in db_categories]
    CAT_LABELS = {**CAT_LABELS_BASE, **{c.key: c.label for c in db_categories}}
    ROLE_LABELS = {**ROLE_LABELS_BASE, **{c.key: c.label for c in db_categories}}

    # セッション形式（設定で追加したカテゴリ）のラベルを取り込む
    sc_row = db.query(AppSetting).filter(AppSetting.key == "session_categories").first()
    try:
        extra_session_cats = json.loads(sc_row.value) if sc_row and sc_row.value else []
    except ValueError:
        extra_session_cats = []
    CAT_LABELS.update({c["key"]: c["label"] for c in extra_session_cats})

    cr_row = db.query(AppSetting).filter(AppSetting.key == "custom_roles").first()
    try:
        custom_roles = json.loads(cr_row.value) if cr_row and cr_row.value else []
    except ValueError:
        custom_roles = []
    ROLE_LABELS.update({r["key"]: r["label"] for r in custom_roles})

    def _cat_header_fill(color_hex: str) -> PatternFill:
        """カテゴリの色からヘッダー用PatternFillを生成"""
        c = color_hex.lstrip("#").upper()
        return PatternFill(start_color=c, end_color=c, fill_type="solid")

    cat_fill_map = {c.key: _cat_header_fill(c.color) for c in db_categories}
    group_label_map = {g.id: g.label for g in db_session_groups}
    room_list = sorted(rooms, key=lambda r: r.id)

    def _is_session_cat(cat: str) -> bool:
        """overall・動的カテゴリ以外は全て通常セッション（部屋列）として扱う"""
        return cat != "overall" and cat not in dynamic_cat_keys

    # ============================================================
    # セッション管理シート (動的カテゴリ除外, 写真付き)
    # ============================================================
    def _fill_mgmt(ws, subset):
        ws.append(["ID", "写真", "タイトル", "登壇者", "ふりがな", "所属", "肩書き", "開始", "終了", "部屋", "カテゴリ", "グループ", "必要人数", "英語", "説明", "備考"])
        _apply_header(ws, 1)
        ws.column_dimensions["B"].width = 9
        for s in subset:
            if not _is_session_cat(s.category):
                continue
            if s.category in ("lt", "panel") and s.lt_talks:
                speakers = "\n".join(
                    f"{t.speaker}（{t.title}）" + (f" {t.start_time}〜{t.end_time}" if t.start_time else "")
                    for t in s.lt_talks
                )
            else:
                speakers = s.speaker
            ws.append([
                s.id,
                "",  # 写真列（画像で上書き）
                s.title,
                speakers,
                s.speaker_kana,
                s.speaker_org,
                s.speaker_title,
                _fmt_full(s.start_time),
                _fmt_full(s.end_time),
                s.room.name if s.room else "",
                CAT_LABELS.get(s.category, s.category),
                group_label_map.get(s.group_id, ""),
                s.required_staff,
                "○" if s.english_required else "",
                s.description,
                s.notes,
            ])
            _apply_border(ws, ws.max_row)
            if s.speaker_photo:
                _add_photo(ws, s.speaker_photo, col=2, row=ws.max_row)
        _auto_width(ws)
        ws.column_dimensions["B"].width = 9  # auto_widthで上書きされるので再設定

    # ============================================================
    # セッション配置シート (写真付き)
    # ============================================================
    def _fill_place(ws, subset):
        ws.append(["時間", "部屋", "写真", "タイトル", "登壇者", "カテゴリ", "必要人数", "配置人数", "英語", "担当スタッフ", "備考"])
        _apply_header(ws, 1, HEADER_FILL_TEAL)
        ws.column_dimensions["C"].width = 9
        for s in subset:
            if not _is_session_cat(s.category):
                continue
            staff_names = ", ".join(a.staff.name for a in s.assignments)
            assigned_count = len(s.assignments)
            status = "○" if assigned_count >= s.required_staff else f"不足({assigned_count}/{s.required_staff})"
            ws.append([
                f"{_fmt(s.start_time)}-{_fmt(s.end_time)}",
                s.room.name if s.room else "",
                "",  # 写真列
                s.title,
                s.speaker,
                CAT_LABELS.get(s.category, s.category),
                s.required_staff,
                status,
                "○" if s.english_required else "",
                staff_names,
                s.notes,
            ])
            _apply_border(ws, ws.max_row)
            if s.speaker_photo:
                _add_photo(ws, s.speaker_photo, col=3, row=ws.max_row)
        _auto_width(ws)
        ws.column_dimensions["C"].width = 9

    # ============================================================
    # 全体スケジュールシート (マトリクス忠実再現)
    # ============================================================
    def _fill_matrix(ws, subset):
        overall_sessions = [s for s in subset if s.category == "overall"]
        session_only = [s for s in subset if _is_session_cat(s.category)]
        cat_sessions_map = {ck: [s for s in subset if s.category == ck] for ck in dynamic_cat_keys}
        all_schedule = overall_sessions + session_only
        for ck in dynamic_cat_keys:
            all_schedule += cat_sessions_map[ck]

        has_overall = len(overall_sessions) > 0
        sess_room_ids = dict.fromkeys(s.room_id for s in session_only if s.room)
        sess_rooms = [r for r in room_list if r.id in sess_room_ids]

        # 列定義: (type, label, fill, room_id_or_None)
        columns = []
        if has_overall:
            columns.append(("overall", "全体", HEADER_FILL_ORANGE, None))
        for r in sess_rooms:
            columns.append(("session", r.name, HEADER_FILL_INDIGO, r.id))
        for cat_obj in db_categories:
            ck = cat_obj.key
            cat_room_ids = dict.fromkeys(s.room_id for s in cat_sessions_map[ck] if s.room)
            cat_rooms = [r for r in room_list if r.id in cat_room_ids]
            fill = cat_fill_map[ck]
            for r in cat_rooms:
                columns.append((ck, f"{cat_obj.label}: {r.name}", fill, r.id))

        if not all_schedule:
            _auto_width(ws)
            return

        # --- 時間スロットを生成 ---
        slot_delta = timedelta(minutes=SLOT_MINUTES)
        min_time = _floor_slot(min(s.start_time for s in all_schedule))
        max_time = _ceil_slot(max(s.end_time for s in all_schedule))
        slots = []
        t = min_time
        while t < max_time:
            slots.append(t)
            t += slot_delta

        HEADER_ROW = 1
        DATA_START = 2

        # --- ヘッダー行 ---
        ws.cell(row=HEADER_ROW, column=1, value="時間")
        ws.cell(row=HEADER_ROW, column=1).font = HEADER_FONT
        ws.cell(row=HEADER_ROW, column=1).fill = HEADER_FILL_INDIGO
        ws.cell(row=HEADER_ROW, column=1).alignment = CENTER
        ws.cell(row=HEADER_ROW, column=1).border = THIN_BORDER

        for ci, (ctype, clabel, cfill, _) in enumerate(columns):
            cell = ws.cell(row=HEADER_ROW, column=ci + 2, value=clabel)
            cell.font = HEADER_FONT
            cell.fill = cfill
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        # --- 時間ラベル (15分ごとに表示) ---
        slot_row_map = {}
        for si, slot_t in enumerate(slots):
            excel_row = DATA_START + si
            slot_row_map[slot_t] = excel_row
            for ci in range(len(columns)):
                cell = ws.cell(row=excel_row, column=ci + 2, value="")
                cell.border = THIN_BORDER
            time_cell = ws.cell(row=excel_row, column=1, value="")
            time_cell.border = THIN_BORDER
            time_cell.alignment = Alignment(horizontal="center", vertical="top")
            if slot_t.minute % 15 == 0:
                time_cell.value = _fmt(slot_t)
                time_cell.font = Font(bold=True, size=9)

        ws.column_dimensions["A"].width = 8
        for si in range(len(slots)):
            ws.row_dimensions[DATA_START + si].height = 14

        # --- カテゴリ別の色定義 ---
        CAT_FILL = {
            "overall": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
            "general": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            "tech": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            "workshop": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            "keynote": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            "lt": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            "panel": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
        }
        CAT_FONT_COLOR = {
            "overall": "E65100",
            "general": "1A73E8", "tech": "1A73E8", "workshop": "1A73E8",
            "keynote": "1A73E8", "lt": "1A73E8", "panel": "1A73E8",
        }
        for cat_obj in db_categories:
            ck = cat_obj.key
            hex_color = cat_obj.color.lstrip("#").upper()
            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
            lr = min(255, r + (255 - r) * 85 // 100)
            lg = min(255, g + (255 - g) * 85 // 100)
            lb = min(255, b + (255 - b) * 85 // 100)
            light_hex = f"{lr:02X}{lg:02X}{lb:02X}"
            CAT_FILL[ck] = PatternFill(start_color=light_hex, end_color=light_hex, fill_type="solid")
            CAT_FONT_COLOR[ck] = hex_color

        # --- セッションをマトリクスに配置 (セル結合) ---
        occupied_cells: set[tuple[int, int]] = set()
        for s in all_schedule:
            cat = s.category
            col_idx = None
            if cat == "overall":
                for ci, (ctype, _, _, _) in enumerate(columns):
                    if ctype == "overall":
                        col_idx = ci + 2
                        break
            elif cat in dynamic_cat_keys:
                for ci, (ctype, _, _, rid) in enumerate(columns):
                    if ctype == cat and rid == s.room_id:
                        col_idx = ci + 2
                        break
            else:  # session categories
                for ci, (ctype, _, _, rid) in enumerate(columns):
                    if ctype == "session" and rid == s.room_id:
                        col_idx = ci + 2
                        break

            if col_idx is None:
                continue

            start_row = None
            end_row = None
            for slot_t, row in slot_row_map.items():
                if slot_t >= s.start_time and start_row is None:
                    start_row = row
                if slot_t < s.end_time:
                    end_row = row

            if start_row is None or end_row is None:
                slot_times_list = sorted(slot_row_map.keys())
                if s.start_time <= slot_times_list[0]:
                    start_row = slot_row_map[slot_times_list[0]]
                else:
                    for st in slot_times_list:
                        if st <= s.start_time:
                            start_row = slot_row_map[st]
                if s.end_time >= slot_times_list[-1]:
                    end_row = slot_row_map[slot_times_list[-1]]
                else:
                    for st in slot_times_list:
                        if st < s.end_time:
                            end_row = slot_row_map[st]

            if start_row is None or end_row is None:
                continue

            staff_names = ", ".join(a.staff.name for a in s.assignments)
            time_str = f"{_fmt(s.start_time)}-{_fmt(s.end_time)}"
            if cat == "overall":
                content = f"{s.title}\n{time_str}"
                if s.required_staff == -1:
                    content += "\n【全員】"
                elif staff_names:
                    content += f"\n【{staff_names}】"
                if s.notes:
                    content += f"\n{s.notes}"
            else:
                content = f"{s.title}\n{time_str}"
                if s.speaker:
                    content += f"\n{s.speaker}"
                if staff_names:
                    content += f"\n【{staff_names}】"
                if not s.assignments and s.required_staff > 0:
                    content += "\n※未配置"

            if (start_row, col_idx) in occupied_cells:
                continue
            for r in range(start_row, end_row + 1):
                occupied_cells.add((r, col_idx))

            if end_row > start_row:
                ws.merge_cells(
                    start_row=start_row, start_column=col_idx,
                    end_row=end_row, end_column=col_idx,
                )

            cell = ws.cell(row=start_row, column=col_idx, value=content)
            fill = CAT_FILL.get(cat, CAT_FILL["general"])
            font_color = CAT_FONT_COLOR.get(cat, "333333")
            cell.fill = fill
            cell.font = Font(size=9, color=font_color)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            cell.border = THIN_BORDER

            if end_row > start_row:
                for r in range(start_row, end_row + 1):
                    border_cell = ws.cell(row=r, column=col_idx)
                    border_cell.border = THIN_BORDER

        for ci, (ctype, clabel, _, _) in enumerate(columns):
            col_letter = get_column_letter(ci + 2)
            label_len = sum(2 if ord(c) > 127 else 1 for c in clabel)
            ws.column_dimensions[col_letter].width = max(label_len + 2, 18)

    # ============================================================
    # 動的カテゴリ別シート
    # ============================================================
    def _fill_cat(ws, cat_obj, subset):
        ws.append(["タイトル", "時間", "場所", "英語", "必要人数", "配置人数", "担当スタッフ", "備考"])
        _apply_header(ws, 1, cat_fill_map[cat_obj.key])
        for s in subset:
            if s.category != cat_obj.key:
                continue
            staff_names = ", ".join(a.staff.name for a in s.assignments)
            assigned_count = len(s.assignments)
            status = "○" if assigned_count >= s.required_staff else f"不足({assigned_count}/{s.required_staff})"
            ws.append([
                s.title,
                f"{_fmt(s.start_time)}-{_fmt(s.end_time)}",
                s.room.name if s.room else "",
                "○" if s.english_required else "",
                s.required_staff,
                status,
                staff_names,
                s.notes,
            ])
            _apply_border(ws, ws.max_row)
        _auto_width(ws)

    # ============================================================
    # ワークブック構築
    # ============================================================
    wb = Workbook()

    _used_sheet_names: set[str] = set()

    def _new_sheet(name):
        """禁止文字・31文字制限に加え、シート名の重複を避けてシートを作成"""
        base = _safe_sheet_name(name)
        title = base
        i = 2
        while title in _used_sheet_names:
            suffix = f" {i}"
            title = base[:31 - len(suffix)] + suffix
            i += 1
        _used_sheet_names.add(title)
        return wb.create_sheet(title)

    # Sheet: 部屋管理
    ws1 = wb.active
    ws1.title = "部屋管理"
    _used_sheet_names.add("部屋管理")
    ws1.append(["ID", "部屋名", "定員", "階"])
    _apply_header(ws1, 1, HEADER_FILL_BROWN)
    for r in rooms:
        if r.name == "全体":
            continue
        ws1.append([r.id, r.name, r.capacity, r.floor])
        _apply_border(ws1, ws1.max_row)
    _auto_width(ws1)

    # Sheet: スタッフ管理
    ws3 = _new_sheet("スタッフ管理")
    ws3.append(["ID", "名前", "Slack名", "担当", "経験回数", "英語", "最大稼働時間", "活動可能時間", "担当セッション数"])
    _apply_header(ws3, 1, HEADER_FILL_ORANGE)
    for st in staffs:
        avails = " / ".join(
            f"{_fmt_md(a.start_time)}-{_fmt(a.end_time)}" for a in st.availabilities
        )
        ws3.append([
            st.id,
            st.name,
            st.slack_name,
            ", ".join(ROLE_LABELS.get(r, r) for r in (st.role or "").split(",") if r) or "なし",
            st.experience_count,
            "○" if st.english_ok else "",
            f"{st.max_hours}h",
            avails,
            len(st.assignments),
        ])
        _apply_border(ws3, ws3.max_row)
    _auto_width(ws3)

    # 日別シート: セッション管理 / セッション配置 / 全体スケジュール / カテゴリ別
    session_dates = sorted({s.start_time.date() for s in sessions})
    multi_day = len(session_dates) > 1

    def _suffix(d):
        return f" {d.strftime('%m-%d')}" if multi_day else ""

    day_list = session_dates or [None]

    def _subset(d):
        return [s for s in sessions if s.start_time.date() == d] if d else sessions

    def _sfx(d):
        return _suffix(d) if d else ""

    # シートは種類ごとにまとめ、各種類内で日別に並べる
    for d in day_list:
        _fill_mgmt(_new_sheet(f"セッション管理{_sfx(d)}"), _subset(d))
    for d in day_list:
        _fill_place(_new_sheet(f"セッション配置{_sfx(d)}"), _subset(d))
    for d in day_list:
        _fill_matrix(_new_sheet(f"全体スケジュール{_sfx(d)}"), _subset(d))
    for cat_obj in db_categories:
        for d in day_list:
            subset = _subset(d)
            if not any(s.category == cat_obj.key for s in subset):
                continue
            _fill_cat(_new_sheet(f"{cat_obj.label}{_sfx(d)}"), cat_obj, subset)

    # Excelファイルをバイトストリームに書き出し
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    ts = app_now().strftime("%Y%m%d_%H%M")
    filename = f"event_schedule_{ts}.xlsx"

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
